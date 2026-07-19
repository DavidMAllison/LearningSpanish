#!/usr/bin/env python3
"""Download YouTube videos from the SI1M Challenge spreadsheet."""

import openpyxl
import subprocess
import os
import json

SPREADSHEET = '/Users/davidallison/Downloads/Video Links - SI1M Challenge Updated.xlsx'
BASE_DIR = '/Users/davidallison/projects/personal/LearningSpanish/LLMContext'
YT_DLP = '/tmp/yt-dlp'

wb = openpyxl.load_workbook(SPREADSHEET)
ws = wb['Sheet1']

non_youtube_links = []
youtube_downloads = []

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False), start=1):
    day_folder = os.path.join(BASE_DIR, f'Day{row_idx}')
    day_label = row[0].value or f'Day{row_idx}'

    for cell in row[1:]:  # Skip column A (day label)
        if cell.value is None or cell.hyperlink is None:
            continue

        name = cell.value.strip().replace('\n', ' ')
        url = cell.hyperlink.target

        if url is None:
            continue

        is_youtube = any(domain in url for domain in ['youtube.com', 'youtu.be'])

        if is_youtube:
            youtube_downloads.append({
                'day': row_idx,
                'day_label': day_label,
                'name': name,
                'url': url,
                'folder': day_folder,
            })
        else:
            non_youtube_links.append({
                'day': row_idx,
                'day_label': day_label,
                'name': name,
                'url': url,
            })

# Save non-youtube links for later
with open(os.path.join(BASE_DIR, 'non_youtube_links.json'), 'w') as f:
    json.dump(non_youtube_links, f, indent=2)

print(f"Total YouTube videos to download: {len(youtube_downloads)}")
print(f"Non-YouTube links saved: {len(non_youtube_links)}")
print()

# Print non-youtube links
print("=== NON-YOUTUBE LINKS (skipped) ===")
for link in non_youtube_links:
    print(f"  Day{link['day']}: {link['name']} -> {link['url']}")
print()

# Download each video
success = 0
failed = []
for i, vid in enumerate(youtube_downloads, 1):
    safe_name = vid['name'].replace('/', '-').replace('"', '').replace("'", "").replace('?', '').replace('!', '').replace(':', ' -')
    output_template = os.path.join(vid['folder'], f"{safe_name}.%(ext)s")

    print(f"[{i}/{len(youtube_downloads)}] Day{vid['day']}: Downloading \"{vid['name']}\"...")
    print(f"  URL: {vid['url']}")
    print(f"  Output: {output_template}")

    try:
        result = subprocess.run(
            [YT_DLP, '-f', 'best[ext=mp4]/best', '--no-playlist',
             '-o', output_template, vid['url']],
            capture_output=True, text=True, timeout=300
        )
        if result.returncode == 0:
            success += 1
            print(f"  SUCCESS")
        else:
            failed.append(vid)
            print(f"  FAILED: {result.stderr[-200:]}")
    except subprocess.TimeoutExpired:
        failed.append(vid)
        print(f"  FAILED: Timeout")
    except Exception as e:
        failed.append(vid)
        print(f"  FAILED: {e}")
    print()

print(f"\n=== SUMMARY ===")
print(f"Downloaded: {success}/{len(youtube_downloads)}")
print(f"Failed: {len(failed)}")
if failed:
    print("\nFailed downloads:")
    for vid in failed:
        print(f"  Day{vid['day']}: {vid['name']} -> {vid['url']}")
